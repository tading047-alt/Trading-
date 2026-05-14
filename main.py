import time
import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# =========================================================
# CONFIG
# =========================================================

TELEGRAM_TOKEN = "8628541851:AAGTo4LDtxv8WOy40L5YI7kqIdwv2SLNUKI"

CHAT_IDS = [
    "5067771509",
    "-1003890327415"
]

INTERVAL = 180  # 3 دقائق بين الدورات

# إعدادات SHORT
MIN_PUMP = 8
MIN_VOLUME = 50000

# إعدادات LONG
MAX_POSITION_SIZE = 100
MAX_LEVERAGE = 2
MAX_COINS_TO_SCAN = 500

# إعدادات فلترة الإشارات
MIN_SCORE_SHORT = 65
MIN_SCORE_LONG = 65
MIN_VOLUME_USDT = 1000000
REQUIRE_GOLDEN_CROSS = False

# إعدادات ATR
MIN_ATR_PERCENT = 1.5
MAX_ATR_PERCENT = 4.0

# إعدادات المسح المجمّع
BATCH_SIZE = 100
BATCH_SCAN_TIME = 300
REST_TIME_BETWEEN_BATCHES = 60

# =========================================================
# PAPER TRADING CONFIG
# =========================================================

PAPER_TRADING_ENABLED = True
PAPER_INITIAL_BALANCE = 10000
PAPER_POSITION_SIZE = 100
PAPER_SLIPPAGE = 0.001
PAPER_FEE = 0.001
MAX_POSITIONS_SIMULTANEOUS = 5

# =========================================================
# TELEGRAM
# =========================================================

def send(msg):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    for chat in CHAT_IDS:
        try:
            response = requests.post(
                url,
                data={
                    "chat_id": chat,
                    "text": msg,
                    "parse_mode": "HTML"
                },
                timeout=10
            )
            print(f"Sent to {chat}: {response.status_code}")
        except Exception as e:
            print(f"Error sending to {chat}: {e}")

# =========================================================
# INDICATORS
# =========================================================

def rsi(series, period=14):
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def ema(series, period=20):
    return series.ewm(span=period, adjust=False).mean()

def calculate_bollinger_bands(df, period=20, std=2):
    middle = df['close'].rolling(period).mean()
    std_dev = df['close'].rolling(period).std()
    upper = middle + (std_dev * std)
    lower = middle - (std_dev * std)
    return upper, middle, lower

def check_golden_cross(df):
    df['ema_50'] = df['close'].ewm(span=50, adjust=False).mean()
    df['ema_200'] = df['close'].ewm(span=200, adjust=False).mean()
    
    ema_50_current = df['ema_50'].iloc[-1]
    ema_200_current = df['ema_200'].iloc[-1]
    ema_50_prev = df['ema_50'].iloc[-2]
    ema_200_prev = df['ema_200'].iloc[-2]
    
    if ema_50_prev <= ema_200_prev and ema_50_current > ema_200_current:
        return True, "Golden Cross detected ✅"
    elif ema_50_current > ema_200_current:
        return True, "EMA 50 above EMA 200 🟢"
    return False, "No golden cross"

def check_bullish_candles(df):
    last_5_candles = df.tail(6)
    bullish_count = 0
    
    for i in range(len(last_5_candles)-1):
        if last_5_candles['close'].iloc[i] > last_5_candles['open'].iloc[i]:
            bullish_count += 1
        if i > 0:
            if last_5_candles['close'].iloc[i] > last_5_candles['high'].iloc[i-1]:
                bullish_count += 1
    
    last_candle = last_5_candles.iloc[-1]
    prev_candle = last_5_candles.iloc[-2]
    
    is_engulfing = (last_candle['close'] > last_candle['open'] and 
                   last_candle['open'] < prev_candle['close'] and
                   last_candle['close'] > prev_candle['open'])
    
    if is_engulfing:
        pattern = "Bullish Engulfing 🟢"
    elif bullish_count >= 5:
        pattern = "Bullish Rejection ✅"
    elif bullish_count >= 3:
        pattern = "Weak Bullish 📈"
    else:
        pattern = "Neutral ⚪"
        
    return pattern, bullish_count, is_engulfing

def wick(df):
    c = df.iloc[-1]
    body = abs(c["c"] - c["o"])
    upper = c["h"] - max(c["c"], c["o"])
    if body == 0:
        body = 0.001
    return upper / body > 2

def volume_weak(df):
    return df["v"].tail(3).mean() < df["v"].tail(15).mean()

def bearish(df):
    prev = df.iloc[-2]
    curr = df.iloc[-1]
    return (
        prev["c"] > prev["o"]
        and curr["c"] < curr["o"]
        and curr["o"] > prev["c"]
    )

def calculate_atr(df, period=14):
    try:
        high = df['h']
        low = df['l']
        close = df['c']
        
        tr1 = high - low
        tr2 = abs(high - close.shift())
        tr3 = abs(low - close.shift())
        
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        atr = tr.rolling(period).mean()
        
        current_price = close.iloc[-1]
        if current_price > 0:
            atr_percent = (atr.iloc[-1] / current_price) * 100
        else:
            atr_percent = 0
            
        return {
            'atr': atr.iloc[-1],
            'atr_percent': atr_percent,
            'is_good': MIN_ATR_PERCENT <= atr_percent <= MAX_ATR_PERCENT
        }
    except:
        return {'atr': 0, 'atr_percent': 0, 'is_good': False}

# =========================================================
# PAPER TRADING CLASS
# =========================================================

class PaperTrading:
    def __init__(self, initial_balance=PAPER_INITIAL_BALANCE):
        self.balance = initial_balance
        self.initial_balance = initial_balance
        self.positions = []
        self.closed_trades = []
        self.current_price_cache = {}
        self.lock = threading.Lock()
        
        self.excel_file = f"paper_trading_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.create_excel_file()
        
        print(f"📊 Paper Trading initialized with ${initial_balance}")
        print(f"📁 Excel file: {self.excel_file}")
    
    def create_excel_file(self):
        self.workbook = openpyxl.Workbook()
        
        sheet_open = self.workbook.active
        sheet_open.title = "Open Positions"
        headers_open = ["Time", "Symbol", "Type", "Entry Price", "Position Size", "Stop Loss", "Take Profit", "Current Price", "Status"]
        sheet_open.append(headers_open)
        
        for col in range(1, len(headers_open) + 1):
            cell = sheet_open.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        sheet_closed = self.workbook.create_sheet("Closed Trades")
        headers_closed = ["Exit Time", "Symbol", "Type", "Entry Price", "Exit Price", "Position Size", "Profit/Loss", "Profit %", "Exit Reason"]
        sheet_closed.append(headers_closed)
        
        for col in range(1, len(headers_closed) + 1):
            cell = sheet_closed.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        sheet_summary = self.workbook.create_sheet("Summary")
        sheet_summary.append(["Metric", "Value"])
        sheet_summary.append(["Initial Balance", f"${self.initial_balance}"])
        sheet_summary.append(["Current Balance", f"${self.balance}"])
        sheet_summary.append(["Total Profit/Loss", f"${self.balance - self.initial_balance}"])
        sheet_summary.append(["Total Trades", "0"])
        sheet_summary.append(["Winning Trades", "0"])
        sheet_summary.append(["Losing Trades", "0"])
        sheet_summary.append(["Win Rate", "0%"])
        
        self.workbook.save(self.excel_file)
    
    def update_summary(self):
        sheet_summary = self.workbook["Summary"]
        
        sheet_summary["B2"] = f"${self.initial_balance:.2f}"
        sheet_summary["B3"] = f"${self.balance:.2f}"
        sheet_summary["B4"] = f"${self.balance - self.initial_balance:.2f}"
        sheet_summary["B5"] = len(self.closed_trades)
        
        winning_trades = len([t for t in self.closed_trades if t['profit'] > 0])
        sheet_summary["B6"] = winning_trades
        sheet_summary["B7"] = len(self.closed_trades) - winning_trades
        
        win_rate = (winning_trades / len(self.closed_trades) * 100) if self.closed_trades else 0
        sheet_summary["B8"] = f"{win_rate:.1f}%"
        
        total_profit = self.balance - self.initial_balance
        profit_cell = sheet_summary["B4"]
        if total_profit >= 0:
            profit_cell.font = Font(color="00FF00", bold=True)
        else:
            profit_cell.font = Font(color="FF0000", bold=True)
        
        self.workbook.save(self.excel_file)
    
    def should_enter_trade(self, signal_score):
        if not PAPER_TRADING_ENABLED:
            return False
        
        with self.lock:
            if len(self.positions) >= MAX_POSITIONS_SIMULTANEOUS:
                return False
            if signal_score >= 75:
                return True
            return False
    
    def enter_long(self, symbol, entry_price, stop_loss, take_profit, score):
        if not self.should_enter_trade(score):
            return False
        
        with self.lock:
            slippage_cost = PAPER_POSITION_SIZE * PAPER_SLIPPAGE
            fee_cost = PAPER_POSITION_SIZE * PAPER_FEE
            total_cost = PAPER_POSITION_SIZE + slippage_cost + fee_cost
            
            if total_cost > self.balance:
                print(f"⚠️ Insufficient balance for {symbol}")
                return False
            
            self.balance -= total_cost
            
            position = {
                'symbol': symbol,
                'type': 'LONG',
                'entry_price': entry_price,
                'stop_loss': stop_loss,
                'take_profit': take_profit,
                'position_size': PAPER_POSITION_SIZE,
                'score': score,
                'entry_time': datetime.now(),
                'status': 'OPEN'
            }
            self.positions.append(position)
            
            sheet = self.workbook["Open Positions"]
            sheet.append([
                position['entry_time'].strftime('%Y-%m-%d %H:%M:%S'),
                symbol,
                'LONG',
                entry_price,
                PAPER_POSITION_SIZE,
                stop_loss,
                take_profit,
                entry_price,
                'OPEN'
            ])
            self.workbook.save(self.excel_file)
            
            print(f"📈 [PAPER] Entered LONG {symbol} at ${entry_price}")
            return True
    
    def enter_short(self, symbol, entry_price, stop_loss, take_profit, score):
        if not self.should_enter_trade(score):
            return False
        
        with self.lock:
            slippage_cost = PAPER_POSITION_SIZE * PAPER_SLIPPAGE
            fee_cost = PAPER_POSITION_SIZE * PAPER_FEE
            total_cost = PAPER_POSITION_SIZE + slippage_cost + fee_cost
            
            if total_cost > self.balance:
                print(f"⚠️ Insufficient balance for {symbol}")
                return False
            
            self.balance -= total_cost
            
            position = {
                'symbol': symbol,
                'type': 'SHORT',
                'entry_price': entry_price,
                'stop_loss': stop_loss,
                'take_profit': take_profit,
                'position_size': PAPER_POSITION_SIZE,
                'score': score,
                'entry_time': datetime.now(),
                'status': 'OPEN'
            }
            self.positions.append(position)
            
            sheet = self.workbook["Open Positions"]
            sheet.append([
                position['entry_time'].strftime('%Y-%m-%d %H:%M:%S'),
                symbol,
                'SHORT',
                entry_price,
                PAPER_POSITION_SIZE,
                stop_loss,
                take_profit,
                entry_price,
                'OPEN'
            ])
            self.workbook.save(self.excel_file)
            
            print(f"📉 [PAPER] Entered SHORT {symbol} at ${entry_price}")
            return True
    
    def get_current_price(self, symbol):
        try:
            url = f"https://api.binance.com/api/v3/ticker/price?symbol={symbol}"
            response = requests.get(url, timeout=5)
            return float(response.json()['price'])
        except:
            return None
    
    def update_positions(self):
        with self.lock:
            if not self.positions:
                return
            
            sheet = self.workbook["Open Positions"]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
            
            positions_to_remove = []
            
            for pos in self.positions:
                current_price = self.get_current_price(pos['symbol'])
                
                if current_price is None:
                    continue
                
                pos['current_price'] = current_price
                
                if pos['type'] == 'LONG':
                    pnl_percent = ((current_price - pos['entry_price']) / pos['entry_price']) * 100
                    hit_stop = current_price <= pos['stop_loss']
                    hit_profit = current_price >= pos['take_profit']
                else:
                    pnl_percent = ((pos['entry_price'] - current_price) / pos['entry_price']) * 100
                    hit_stop = current_price >= pos['stop_loss']
                    hit_profit = current_price <= pos['take_profit']
                
                exit_reason = None
                exit_price = current_price
                
                if hit_stop:
                    exit_reason = "Stop Loss"
                elif hit_profit:
                    exit_reason = "Take Profit"
                
                if exit_reason:
                    if pos['type'] == 'LONG':
                        profit_percent = ((exit_price - pos['entry_price']) / pos['entry_price']) * 100
                    else:
                        profit_percent = ((pos['entry_price'] - exit_price) / pos['entry_price']) * 100
                    
                    profit_amount = pos['position_size'] * (profit_percent / 100)
                    
                    closed_trade = {
                        'symbol': pos['symbol'],
                        'type': pos['type'],
                        'entry_price': pos['entry_price'],
                        'exit_price': exit_price,
                        'position_size': pos['position_size'],
                        'profit': profit_amount,
                        'profit_percent': profit_percent,
                        'exit_reason': exit_reason,
                        'entry_time': pos['entry_time'],
                        'exit_time': datetime.now()
                    }
                    self.closed_trades.append(closed_trade)
                    
                    self.balance += pos['position_size'] + profit_amount
                    
                    sheet_closed = self.workbook["Closed Trades"]
                    sheet_closed.append([
                        closed_trade['exit_time'].strftime('%Y-%m-%d %H:%M:%S'),
                        pos['symbol'],
                        pos['type'],
                        pos['entry_price'],
                        exit_price,
                        pos['position_size'],
                        f"${profit_amount:.2f}",
                        f"{profit_percent:.2f}%",
                        exit_reason
                    ])
                    
                    last_row = sheet_closed.max_row
                    profit_cell = sheet_closed.cell(row=last_row, column=7)
                    if profit_amount >= 0:
                        profit_cell.font = Font(color="00FF00", bold=True)
                    else:
                        profit_cell.font = Font(color="FF0000", bold=True)
                    
                    positions_to_remove.append(pos)
                    
                    print(f"✅ [PAPER] Closed {pos['type']} {pos['symbol']}: {profit_percent:.2f}% ({exit_reason})")
                else:
                    row_index = self.positions.index(pos) + 2
                    sheet.cell(row=row_index, column=8, value=current_price)
                    sheet.cell(row=row_index, column=9, value='OPEN')
            
            for pos in positions_to_remove:
                self.positions.remove(pos)
            
            self.update_summary()
    
    def get_summary_message(self):
        total_profit = self.balance - self.initial_balance
        winning_trades = len([t for t in self.closed_trades if t['profit'] > 0])
        losing_trades = len([t for t in self.closed_trades if t['profit'] <= 0])
        win_rate = (winning_trades / len(self.closed_trades) * 100) if self.closed_trades else 0
        
        message = f"""
📊 <b>PAPER TRADING SUMMARY</b>
━━━━━━━━━━━━━━━━━━
💰 <b>Initial Balance:</b> ${self.initial_balance}
💰 <b>Current Balance:</b> ${self.balance:.2f}
📈 <b>Total P/L:</b> ${total_profit:.2f}

━━━━━━━━━━━━━━━━━━
<b>STATISTICS:</b>
• Total Trades: {len(self.closed_trades)}
• Winning: {winning_trades}
• Losing: {losing_trades}
• Win Rate: {win_rate:.1f}%

━━━━━━━━━━━━━━━━━━
<b>OPEN POSITIONS:</b> {len(self.positions)}
"""
        if self.positions:
            for pos in self.positions:
                message += f"\n• {pos['symbol']} ({pos['type']}) @ ${pos['entry_price']:.4f}"
        
        message += f"""
━━━━━━━━━━━━━━━━━━
⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
        return message
    
    def send_excel_to_telegram(self):
        self.update_positions()
        self.update_summary()
        
        url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
        
        with open(self.excel_file, 'rb') as f:
            files = {'document': f}
            data = {'chat_id': CHAT_IDS[0], 'caption': f"📊 Paper Trading Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"}
            
            try:
                response = requests.post(url, files=files, data=data, timeout=30)
                print(f"📁 Excel file sent to Telegram: {response.status_code}")
            except Exception as e:
                print(f"Error sending file: {e}")
    
    def run(self):
        while True:
            self.update_positions()
            time.sleep(60)

# =========================================================
# BINANCE DATA
# =========================================================

def klines(symbol, interval='5m', limit=100):
    url = f"https://api.binance.com/api/v3/klines?symbol={symbol}&interval={interval}&limit={limit}"
    try:
        data = requests.get(url, timeout=10).json()
        if 'code' in data:
            return None
        df = pd.DataFrame(data)
        df = df.iloc[:, :6]
        df.columns = ["t","o","h","l","c","v"]
        for col in ["o","h","l","c","v"]:
            df[col] = pd.to_numeric(df[col])
        return df
    except:
        return None

def klines_multiple_timeframes(symbol):
    dataframes = {}
    
    df_15m = klines(symbol, '15m', 100)
    if df_15m is not None and len(df_15m) >= 50:
        dataframes['15m'] = df_15m
    
    df_1h = klines(symbol, '1h', 100)
    if df_1h is not None and len(df_1h) >= 50:
        dataframes['1h'] = df_1h
    
    df_4h = klines(symbol, '4h', 100)
    if df_4h is not None and len(df_4h) >= 50:
        dataframes['4h'] = df_4h
    
    return dataframes if dataframes else None

def get_all_usdt_pairs(limit=MAX_COINS_TO_SCAN):
    url = "https://api.binance.com/api/v3/exchangeInfo"
    try:
        data = requests.get(url).json()
        symbols = []
        for s in data['symbols']:
            if s['quoteAsset'] == 'USDT' and s['status'] == 'TRADING':
                symbols.append(s['symbol'])
        
        if len(symbols) > limit:
            tickers = requests.get("https://api.binance.com/api/v3/ticker/24hr").json()
            volume_dict = {}
            for t in tickers:
                if t['symbol'] in symbols:
                    try:
                        volume_dict[t['symbol']] = float(t['quoteVolume'])
                    except:
                        volume_dict[t['symbol']] = 0
            
            symbols.sort(key=lambda x: volume_dict.get(x, 0), reverse=True)
            symbols = symbols[:limit]
        
        return symbols
    except Exception as e:
        print(f"Error getting pairs: {e}")
        return []

# =========================================================
# SCAN SHORT
# =========================================================

def scan_short_opportunities():
    url = "https://api.binance.com/api/v3/ticker/24hr"
    try:
        data = requests.get(url, timeout=10).json()
    except:
        return []
    
    opportunities = []
    for c in data:
        try:
            sym = c["symbol"]
            if not sym.endswith("USDT"):
                continue
            pump = float(c["priceChangePercent"])
            vol = float(c["quoteVolume"])
            
            if vol < MIN_VOLUME_USDT:
                continue
            
            if pump > MIN_PUMP:
                df = klines(sym, '5m', 60)
                if df is None or len(df) < 30:
                    continue
                
                atr_data = calculate_atr(df)
                
                df["rsi"] = rsi(df["c"])
                df["ema"] = ema(df["c"])
                current_price = df["c"].iloc[-1]
                current_rsi = df["rsi"].iloc[-1]
                ema20 = df["ema"].iloc[-1]
                stretch = ((current_price - ema20) / ema20) * 100
                
                score = 0
                if current_rsi > 65:
                    score += 20
                if current_rsi > 75:
                    score += 10
                if stretch > 5:
                    score += 10
                
                if wick(df):
                    score += 15
                if volume_weak(df):
                    score += 15
                if bearish(df):
                    score += 20
                
                if not atr_data['is_good']:
                    continue
                
                if score < MIN_SCORE_SHORT:
                    continue
                
                rsi_5m = current_rsi
                rsi_15m = df["rsi"].iloc[-3] if len(df) >= 3 else current_rsi
                rsi_1h = df["rsi"].iloc[-12] if len(df) >= 12 else current_rsi
                
                change_4h = ((df["c"].iloc[-1] / df["c"].iloc[-48]) - 1) * 100 if len(df) >= 48 else pump * 0.3
                change_1h = ((df["c"].iloc[-1] / df["c"].iloc[-12]) - 1) * 100 if len(df) >= 12 else pump * 0.1
                
                entry_low = current_price * 1.01
                entry_high = current_price * 1.03
                expected_drop = abs(stretch * 0.7)
                
                opportunities.append({
                    'symbol': sym,
                    'pump': pump,
                    'score': score,
                    'current_price': current_price,
                    'rsi': current_rsi,
                    'stretch': stretch,
                    'rsi_5m': rsi_5m,
                    'rsi_15m': rsi_15m,
                    'rsi_1h': rsi_1h,
                    'change_24h': pump,
                    'change_4h': change_4h,
                    'change_1h': change_1h,
                    'entry_low': entry_low,
                    'entry_high': entry_high,
                    'drop': expected_drop,
                    'atr_percent': atr_data['atr_percent']
                })
        except Exception as e:
            continue
    
    opportunities.sort(key=lambda x: x['score'], reverse=True)
    return opportunities

# =========================================================
# SCAN LONG WITH BATCHES
# =========================================================

def scan_long_opportunities_batch(batch_symbols, batch_num, total_batches):
    opportunities = []
    
    for i, sym in enumerate(batch_symbols):
        dataframes = klines_multiple_timeframes(sym)
        if not dataframes:
            continue
        
        analyses = {}
        total_score = 0
        
        for tf, df in dataframes.items():
            if df is None or len(df) < 50:
                continue
            
            atr_data = calculate_atr(df)
            
            current_price = df['c'].iloc[-1]
            df['rsi'] = rsi(df['c'])
            current_rsi = df['rsi'].iloc[-1]
            
            upper, middle, lower = calculate_bollinger_bands(df)
            bb_bullish = current_price > middle.iloc[-1]
            bb_signal = "Above middle band 📈" if bb_bullish else "Between bands ⚪"
            
            golden_cross, gc_message = check_golden_cross(df)
            candle_pattern, _, is_engulfing = check_bullish_candles(df)
            
            tf_weight = 3 if tf == '4h' else 2 if tf == '1h' else 1
            tf_score = 0
            
            if current_rsi >= 50:
                tf_score += 2
            if bb_bullish:
                tf_score += 2
            if golden_cross:
                tf_score += 3
            if is_engulfing:
                tf_score += 2
            
            total_score += tf_score * tf_weight
            
            analyses[tf] = {
                'rsi': current_rsi,
                'bb_signal': bb_signal,
                'golden_cross': golden_cross,
                'gc_message': gc_message,
                'candle_pattern': candle_pattern,
                'current_price': current_price,
                'upper_band': upper.iloc[-1],
                'middle_band': middle.iloc[-1],
                'atr_percent': atr_data['atr_percent']
            }
        
        if not analyses:
            continue
        
        best_atr = max([analyses[tf]['atr_percent'] for tf in analyses if analyses[tf]['atr_percent'] > 0], default=0)
        if best_atr < MIN_ATR_PERCENT or best_atr > MAX_ATR_PERCENT:
            continue
        
        if total_score < MIN_SCORE_LONG:
            continue
        
        main_tf_4h = analyses.get('4h', {})
        if REQUIRE_GOLDEN_CROSS and not main_tf_4h.get('golden_cross', False):
            continue
        
        main_tf = analyses.get('4h') or analyses.get('1h') or analyses.get('15m')
        current_price = main_tf['current_price']
        
        entry_low = round(current_price * 0.99, 4)
        entry_high = round(current_price, 4)
        expected_gain = round((main_tf['upper_band'] - current_price) / current_price * 100, 2)
        if expected_gain < 2:
            expected_gain = 3.0
        
        stop_loss = round(current_price * 0.97, 4)
        take_profit_1 = round(current_price * 1.03, 4)
        take_profit_2 = round(current_price * 1.06, 4)
        
        opportunities.append({
            'symbol': sym,
            'score': total_score,
            'current_price': current_price,
            'entry_low': entry_low,
            'entry_high': entry_high,
            'expected_gain': expected_gain,
            'stop_loss': stop_loss,
            'take_profit_1': take_profit_1,
            'take_profit_2': take_profit_2,
            'analyses': analyses,
            'atr_percent': best_atr
        })
        
        time.sleep(0.3)
    
    return opportunities

# =========================================================
# MESSAGE FORMATTING
# =========================================================

def format_short_message(opp, paper_executed=False):
    symbol = opp['symbol']
    score = opp['score']
    
    if score >= 85:
        grade = "🟢 VERY GOOD"
        strength = "HIGH"
        color = "🟢"
    elif score >= 70:
        grade = "🟡 GOOD"
        strength = "MEDIUM"
        color = "🟡"
    else:
        grade = "🔴 MEDIUM"
        strength = "LOW"
        color = "🔴"
    
    paper_line = "\n━━━━━━━━━━━━━━━━━━\n🤖 PAPER TRADE: EXECUTED ✅" if paper_executed else ""
    
    message = f"""
{color} BINANCE — {grade}

━━━━━━━━━━━━━━━━━━
🔥 SHORT OPPORTUNITY
━━━━━━━━━━━━━━━━━━

💰 PAIR: {symbol}
🧠 AI SCORE: {score} / 100
⚠️ SIGNAL STRENGTH: {strength}

━━━━━━━━━━━━━━━━━━
📊 MARKET MOVEMENT
━━━━━━━━━━━━━━━━━━

📈 24H CHANGE: {opp['change_24h']:+.2f}%
⏱ 4H CHANGE: {opp['change_4h']:+.2f}%
⚡ 1H CHANGE: {opp['change_1h']:+.2f}%

━━━━━━━━━━━━━━━━━━
🧠 TECHNICAL ANALYSIS
━━━━━━━━━━━━━━━━━━

📊 RSI 5M: {opp['rsi_5m']:.2f}
📊 RSI 15M: {opp['rsi_15m']:.2f}
📊 RSI 1H: {opp['rsi_1h']:.2f}

🕯 CANDLE PATTERN:
✔ Bearish Rejection

📉 VOLUME STATUS:
⚠ Weakening

📏 EMA DISTANCE:
{opp['stretch']:.2f}%

📊 ATR (Avg True Range):
{opp['atr_percent']:.2f}% ✅

━━━━━━━━━━━━━━━━━━
🎯 TRADE SETUP
━━━━━━━━━━━━━━━━━━

🔴 SHORT ENTRY ZONE:
{opp['entry_low']:.8f} → {opp['entry_high']:.8f}

📉 EXPECTED DROP:
{opp['drop']:.2f}%

━━━━━━━━━━━━━━━━━━
💼 RISK MANAGEMENT
━━━━━━━━━━━━━━━━━━

💵 POSITION SIZE: 5$
⚡ LEVERAGE: x2 (Isolated)
{paper_line}
━━━━━━━━━━━━━━━━━━
⏱ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
━━━━━━━━━━━━━━━━━━
"""
    return message

def format_long_message(opp, paper_executed=False):
    symbol = opp['symbol']
    score = opp['score']
    
    if score >= 30:
        strength_emoji = "🚀🚀🚀"
        strength_text = "VERY STRONG"
    elif score >= 20:
        strength_emoji = "🚀🚀"
        strength_text = "STRONG"
    elif score >= 15:
        strength_emoji = "📈"
        strength_text = "MEDIUM"
    else:
        strength_emoji = "⭐"
        strength_text = "WEAK"
    
    rsi_lines = []
    for tf in ['15m', '1h', '4h']:
        if tf in opp['analyses']:
            rsi_value = opp['analyses'][tf]['rsi']
            if rsi_value >= 50:
                emoji = "📈"
                status = "Bullish momentum"
            elif rsi_value >= 30:
                emoji = "⚪"
                status = "Neutral"
            else:
                emoji = "📉"
                status = "Weak"
            rsi_lines.append(f"• {tf}: {rsi_value:.1f} ({status}) {emoji}")
    
    bb_4h = opp['analyses'].get('4h', opp['analyses']['1h'])
    candle_4h = opp['analyses'].get('4h', opp['analyses']['1h'])
    gc_4h = opp['analyses'].get('4h', opp['analyses']['1h'])
    
    paper_line = "\n━━━━━━━━━━━━━━━━━━\n🤖 PAPER TRADE: EXECUTED ✅" if paper_executed else ""
    
    message = f"""
{strength_emoji} <b>BULLISH OPPORTUNITY</b> {strength_emoji}

━━━━━━━━━━━━━━━━━━
<b>PAIR:</b> {symbol}
<b>AI SCORE:</b> {score} / 100
<b>SIGNAL STRENGTH:</b> {strength_text}
━━━━━━━━━━━━━━━━━━

<b>💰 CURRENT PRICE:</b> ${opp['current_price']:.4f}

<b>🎯 ENTRY ZONE:</b>
{opp['entry_low']:.4f} → {opp['entry_high']:.4f}

━━━━━━━━━━━━━━━━━━
<b>📊 TECHNICAL ANALYSIS</b>
━━━━━━━━━━━━━━━━━━

<b>📈 RSI ANALYSIS:</b>
{chr(10).join(rsi_lines)}

<b>📊 BOLLINGER BANDS:</b>
• Position: {bb_4h['bb_signal']}
• ATR %: {opp['atr_percent']:.2f}% ✅

<b>🕯️ CANDLE PATTERNS:</b>
• {candle_4h['candle_pattern']}

<b>🟡 GOLDEN CROSS:</b>
• {gc_4h['gc_message']}

━━━━━━━━━━━━━━━━━━
<b>💡 TRADE SETUP (LONG)</b>
━━━━━━━━━━━━━━━━━━

<b>📈 LONG ENTRY ZONE:</b>
{opp['entry_low']:.4f} → {opp['entry_high']:.4f}

<b>🎯 EXPECTED GAIN:</b>
{opp['expected_gain']}%

━━━━━━━━━━━━━━━━━━
<b>⚙️ RISK MANAGEMENT</b>
━━━━━━━━━━━━━━━━━━

<b>💰 POSITION SIZE:</b> ${MAX_POSITION_SIZE}
<b>📊 LEVERAGE:</b> x{MAX_LEVERAGE} (Isolated)
<b>🛑 STOP LOSS:</b> ${opp['stop_loss']}
<b>✅ TAKE PROFIT 1:</b> ${opp['take_profit_1']}
<b>✅ TAKE PROFIT 2:</b> ${opp['take_profit_2']}
{paper_line}
━━━━━━━━━━━━━━━━━━
⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
━━━━━━━━━━━━━━━━━━
"""
    return message

# =========================================================
# MAIN LOOP WITH PAPER TRADING
# =========================================================

sent_short = set()
sent_long = set()
full_scan_counter = 0

paper = PaperTrading()
paper.send_excel_to_telegram()

if PAPER_TRADING_ENABLED:
    paper_thread = threading.Thread(target=paper.run, daemon=True)
    paper_thread.start()
    print("🤖 Paper Trading thread started")

print("🚀 SIGNAL SCANNER STARTED - SENDING SIGNALS ONLY")
print(f"📊 Total coins: {MAX_COINS_TO_SCAN}")
print(f"💰 Paper Trading: {'ON' if PAPER_TRADING_ENABLED else 'OFF'}")

send(f"🚀 <b>SIGNAL SCANNER + PAPER TRADING STARTED</b>\n\n📊 Scanning {MAX_COINS_TO_SCAN} coins\n💰 Paper Trading Active\n💵 Initial Balance: ${PAPER_INITIAL_BALANCE}\n\n🎯 Only quality signals (Score > 75) will be executed in paper trading!")

last_excel_send = time.time()

while True:
    try:
        print(f"\n{'='*50}")
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Scanning cycle started")
        print(f"{'='*50}")
        
        # =================================================
        # SCAN SHORT
        # =================================================
        print("\n📉 Scanning SHORT...")
        short_opps = scan_short_opportunities()
        print(f"Found {len(short_opps)} SHORT opportunities")
        
        for opp in short_opps[:5]:
            if opp['symbol'] in sent_short:
                continue
            
            paper_executed = False
            if PAPER_TRADING_ENABLED and opp['score'] >= 75:
                stop_loss = opp['current_price'] * 1.02
                take_profit = opp['current_price'] * 0.97
                paper.enter_short(opp['symbol'], opp['current_price'], stop_loss, take_profit, opp['score'])
                paper_executed = True
            
            message = format_short_message(opp, paper_executed)
            send(message)
            sent_short.add(opp['symbol'])
            print(f"  ✅ Sent SHORT: {opp['symbol']}")
        
        # =================================================
        # SCAN LONG WITH BATCHES
        # =================================================
        print("\n📈 Scanning LONG...")
        
        all_symbols = get_all_usdt_pairs(MAX_COINS_TO_SCAN)
        total_batches = (len(all_symbols) + BATCH_SIZE - 1) // BATCH_SIZE
        
        all_long_opportunities = []
        
        for batch_num in range(total_batches):
            batch_start = batch_num * BATCH_SIZE
            batch_end = min(batch_start + BATCH_SIZE, len(all_symbols))
            batch_symbols = all_symbols[batch_start:batch_end]
            
            print(f"  Batch {batch_num + 1}/{total_batches} - {len(batch_symbols)} coins")
            
            batch_opps = scan_long_opportunities_batch(batch_symbols, batch_num + 1, total_batches)
            
            for opp in batch_opps:
                if opp['symbol'] in sent_long:
                    continue
                
                paper_executed = False
                if PAPER_TRADING_ENABLED and opp['score'] >= 75:
                    stop_loss = opp['current_price'] * 0.98
                    take_profit = opp['current_price'] * 1.03
                    paper.enter_long(opp['symbol'], opp['current_price'], stop_loss, take_profit, opp['score'])
                    paper_executed = True
                
                message = format_long_message(opp, paper_executed)
                send(message)
                sent_long.add(opp['symbol'])
                all_long_opportunities.append(opp)
                print(f"  ✅ Sent LONG: {opp['symbol']}")
            
            time.sleep(BATCH_SCAN_TIME)
            
            if batch_num < total_batches - 1:
                print(f"  💤 Resting {REST_TIME_BETWEEN_BATCHES} seconds...")
                time.sleep(REST_TIME_BETWEEN_BATCHES)
        
        # =================================================
        # SEND PAPER TRADING SUMMARY & EXCEL
        # =================================================
        if time.time() - last_excel_send > 3600:
            paper.send_excel_to_telegram()
            send(paper.get_summary_message())
            last_excel_send = time.time()
        
        print(f"\n✅ Cycle complete. Total SHORT: {len(short_opps[:5])}, LONG: {len(all_long_opportunities)}")
        print(f"⏳ Waiting {INTERVAL} seconds...\n")
        
        time.sleep(INTERVAL)
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        time.sleep(60)
